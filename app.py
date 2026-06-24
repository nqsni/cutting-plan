from flask import Flask, request, jsonify, render_template, send_from_directory
import os, re
from openpyxl import load_workbook
from engine.ffd import run_ffd_1d, run_ffd_2d

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def normalize_number(val):
    if val is None:
        return None
    s = str(val).strip().replace(' ', '')
    if not s:
        return None
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def parse_spec_v2(spec_raw):
    """
    Parse format Profile versi 2:
      P     105x20  → PLAT, lebar=105, tebal=20 (group key = tebal)
      L  45x 45x 5  → PROFIL siku (group key = spec lengkap)
      R  16x 4      → PROFIL (group key = spec lengkap)
    Aturan: angka TERAKHIR setelah x/X terakhir = tebal (untuk PLAT)
    """
    s = str(spec_raw or '').strip()
    s_clean = re.sub(r'\s+', ' ', s)  # normalisasi spasi berlebih

    # PLAT: diawali P (diikuti spasi dan angka, lalu x, lalu angka)
    # Contoh: P     105x20, P     170x16
    plat_match = re.match(r'^P\s+([\d.,]+)\s*[xX]\s*([\d.,]+)$', s_clean, re.I)
    if plat_match:
        lebar = normalize_number(plat_match.group(1))
        tebal = normalize_number(plat_match.group(2))
        return {
            'kategori':  'PLAT',
            'group_key': f'P_{tebal}',           # grup per tebal
            'label':     f'P tebal {_fmt(tebal)} mm',
            'dimensi':   {'tebal': tebal, 'lebar': lebar},
            'bisa_cut':  True,
        }

    # PROFIL: L, R, H, dll — group key = spec lengkap (sudah dinormalisasi)
    norm_key = re.sub(r'\s+', '', s_clean).upper()  # hapus semua spasi
    return {
        'kategori':  'PROFIL',
        'group_key': norm_key,
        'label':     s_clean,
        'dimensi':   {},
        'bisa_cut':  True,
    }


def _fmt(n):
    if n is None: return '?'
    return str(int(n)) if n == int(n) else str(n)


def parse_excel_boq_v2(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Cari header row: cari kolom 'Mark'
    header_row = 1
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            if str(cell.value or '').strip().lower() == 'mark':
                header_row = cell.row
                break
        else:
            continue
        break

    # Map kolom dari header
    col_map = {}
    for cell in ws[header_row]:
        v = str(cell.value or '').strip().lower().replace(' ', '')
        col = cell.column - 1
        if v == 'no.' or v == 'no':
            col_map['no'] = col
        elif v == 'mark':
            col_map['mark'] = col
        elif v == 'submark':
            col_map['submark'] = col
        elif v == 'qty':
            col_map['qty'] = col
        elif v == 'profile':
            col_map['spesifikasi'] = col
        elif v == 'mtl' or v == 'material':
            col_map['mtl'] = col
        elif v == 'length':
            col_map['panjang'] = col
        elif v == 'weight':
            col_map['berat'] = col
        elif v == 'remark':
            col_map['remark'] = col

    # Fallback: No|Mark|Sub Mark|QTY|Profile|Mtl|Length|Weight|Total Weight|Remark
    defaults = {'no': 0, 'mark': 1, 'submark': 2, 'qty': 3,
                'spesifikasi': 4, 'mtl': 5, 'panjang': 6, 'berat': 7}
    for k, v in defaults.items():
        if k not in col_map:
            col_map[k] = v

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue

        def gcol(key):
            idx = col_map.get(key, -1)
            if idx < 0 or idx >= len(row): return None
            return row[idx]

        mark    = str(gcol('mark') or '').strip()
        submark = str(gcol('submark') or '').strip()
        if not mark or mark.lower() in ('mark', '-', ''):
            continue

        # Drawing ID = Mark + Sub Mark (kalau ada)
        drawing_no = f"{mark} {submark}".strip() if submark else mark
        partname   = submark or mark

        spec_raw = str(gcol('spesifikasi') or '').strip()
        panjang  = normalize_number(gcol('panjang'))
        mtl      = str(gcol('mtl') or '').strip()
        qty      = int(normalize_number(gcol('qty')) or 1)

        if not spec_raw:
            continue

        if panjang is None or panjang <= 0:
            spec_info = {
                'kategori':  'PCS',
                'group_key': spec_raw.upper().strip(),
                'label':     spec_raw,
                'dimensi':   {},
                'bisa_cut':  False,
            }
        else:
            spec_info = parse_spec_v2(spec_raw)

        items.append({
            'drawing_no': drawing_no,
            'partname':   partname,
            'spec_raw':   spec_raw,
            'spec_info':  spec_info,
            'panjang':    panjang,
            'mtl':        mtl,
            'qty':        qty,
        })

    return items


def group_items(items):
    groups = {}
    for item in items:
        si  = item['spec_info']
        key = si.get('group_key') or item['spec_raw'].upper().strip()
        if key not in groups:
            groups[key] = {
                'spec_key':  key,
                'spec_raw':  item['spec_raw'],
                'spec_info': si,
                'items':     [],
                'total_qty': 0,
            }
        groups[key]['items'].append(item)
        groups[key]['total_qty'] += item['qty']
    return list(groups.values())


def _build_group_response(g):
    si       = g['spec_info']
    kategori = si.get('kategori', 'UNKNOWN')
    grp = {
        'spec_key':    g['spec_key'],
        'spec_raw':    g['spec_raw'],
        'kategori':    kategori,
        'label':       si.get('label', g['spec_raw']),
        'bisa_cut':    si.get('bisa_cut', False),
        'total_qty':   g['total_qty'],
        'items_count': len(g['items']),
        'items': [
            {
                'drawing_no': it['drawing_no'],
                'partname':   it['partname'],
                'spec_raw':   it['spec_raw'],
                'lebar':      it['spec_info'].get('dimensi', {}).get('lebar'),
                'panjang':    it['panjang'],
                'mtl':        it['mtl'],
                'qty':        it['qty'],
            }
            for it in g['items']
        ],
    }
    if si.get('bisa_cut') and kategori == 'PLAT':
        lebarset = sorted(set(
            it['spec_info'].get('dimensi', {}).get('lebar')
            for it in g['items']
            if it['spec_info'].get('dimensi', {}).get('lebar') is not None
        ))
        grp['dimensi'] = {
            'type': '2D',
            'tebal': si.get('dimensi', {}).get('tebal'),
            'lebar_variants': lebarset,
        }
    elif si.get('bisa_cut'):
        grp['dimensi'] = {'type': '1D'}
    return grp


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/parse-excel', methods=['POST'])
def parse_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang diupload'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format file harus .xlsx atau .xls'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, 'boq_temp.xlsx')
    f.save(filepath)
    try:
        items = parse_excel_boq_v2(filepath)
    except Exception as e:
        return jsonify({'error': f'Gagal membaca Excel: {str(e)}'}), 500

    if not items:
        return jsonify({'error': 'Tidak ada data valid'}), 400

    groups = group_items(items)
    return jsonify({
        'groups':       [_build_group_response(g) for g in groups],
        'total_items':  len(items),
        'total_groups': len(groups),
    })


@app.route('/api/calculate', methods=['POST'])
def calculate():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Data tidak valid'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, 'boq_temp.xlsx')
    if not os.path.exists(filepath):
        return jsonify({'error': 'File tidak ditemukan, silakan upload ulang'}), 400

    try:
        items = parse_excel_boq_v2(filepath)
    except Exception as e:
        return jsonify({'error': f'Gagal membaca Excel: {str(e)}'}), 500

    groups      = group_items(items)
    raw_configs = data.get('raw_configs', {})
    results, eff_list = [], []
    summary = {'total_raw_material': 0, 'total_groups': len(groups),
                'cuttable_groups': 0, 'pcs_groups': 0}

    for g in groups:
        si       = g['spec_info']
        spec_key = g['spec_key']
        kategori = si.get('kategori', 'UNKNOWN')
        grp_result = _build_group_response(g)

        if not si.get('bisa_cut'):
            grp_result['mode'] = 'PCS'
            summary['pcs_groups'] += 1
            results.append(grp_result)
            continue

        summary['cuttable_groups'] += 1
        cfg = raw_configs.get(spec_key, {})

        if kategori == 'PLAT':
            raw_w, raw_h, kerf = float(cfg.get('raw_w', 1200)), float(cfg.get('raw_h', 2400)), 5
            pieces = []
            for it in g['items']:
                if not it['panjang'] or it['panjang'] <= 0: continue
                lebar_pot = it['spec_info'].get('dimensi', {}).get('lebar') or raw_w
                for _ in range(it['qty']):
                    pieces.append({'w': float(lebar_pot), 'h': float(it['panjang']),
                                   'drawing_no': it['drawing_no'], 'label': it['drawing_no']})
            if not pieces:
                grp_result['mode'] = 'PCS'; results.append(grp_result); continue

            sheets     = run_ffd_2d(pieces, raw_w, raw_h, kerf)
            area_used  = sum(p['w'] * p['h'] for p in pieces)
            efficiency = round(area_used / (len(sheets) * raw_w * raw_h) * 100, 1)
            grp_result.update({'mode': '2D', 'raw_w': raw_w, 'raw_h': raw_h,
                               'sheets': sheets, 'num_sheets': len(sheets),
                               'efficiency': efficiency, 'pieces_count': len(pieces)})
            eff_list.append(efficiency)
            summary['total_raw_material'] += len(sheets)

        else:
            raw_length = float(cfg.get('raw_length', 12000))
            pieces = []
            for it in g['items']:
                if not it['panjang'] or it['panjang'] <= 0: continue
                for _ in range(it['qty']):
                    pieces.append({'length': float(it['panjang']),
                                   'drawing_no': it['drawing_no'], 'label': it['drawing_no']})
            if not pieces:
                grp_result['mode'] = 'PCS'; results.append(grp_result); continue

            bars       = run_ffd_1d(pieces, raw_length, 0)
            len_used   = sum(p['length'] for p in pieces)
            efficiency = round(len_used / (len(bars) * raw_length) * 100, 1)
            grp_result.update({'mode': '1D', 'raw_length': raw_length,
                               'bars': bars, 'num_bars': len(bars),
                               'efficiency': efficiency, 'pieces_count': len(pieces)})
            eff_list.append(efficiency)
            summary['total_raw_material'] += len(bars)

        results.append(grp_result)

    summary['avg_efficiency'] = round(sum(eff_list) / len(eff_list), 1) if eff_list else 0
    return jsonify({'results': results, 'summary': summary})


@app.route("/download-template")
def download_template():
    return send_from_directory("uploads", "template_upload_BQ.xlsx", as_attachment=True)


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5050, debug=True)